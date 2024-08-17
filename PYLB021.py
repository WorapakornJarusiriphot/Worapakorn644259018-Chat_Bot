from flask import Flask, request
from dotenv import load_dotenv
import os
from linebot import LineBotApi, WebhookHandler
from linebot.models import (MessageEvent,
                            TextMessage,
                            TextSendMessage)

from openpyxl import load_workbook

line_bot_api = None
handler = None

def init(secret, token):
    global line_bot_api, handler
    line_bot_api = LineBotApi(token)
    handler = WebhookHandler(secret)

app = Flask(__name__)

@app.route("/", methods=["GET","POST"])
def home():
    try:
        signature = request.headers["X-Line-Signature"]
        body = request.get_data(as_text=True)
        handler.handle(body, signature)
    except:
        pass
    
    return "Hello Line Chatbot"


def handle_text_message(event):
    text = event.message.text
    print(text)

    wb = load_workbook(filename='product.xlsx', read_only=True)
    ws = wb['Sheet1']

    if (text.startswith("ค้นหา")): # "ค้นหา TV7788"
        text_id = text.strip("ค้นหา").replace(" ", "") # "TV7788"
        print(text_id) # "TV7788"
        
    for i in range(2,ws.max_row+1):              # เริ่มอ่านที่ cell A2 เป็นต้นไป
        if text_id == str(ws["A"+str(i)].value): # ค้นหารหัสสินค้า
            product_name = str(ws["B"+str(i)].value) # ชื่อสินค้า
            product_price = str(ws["C"+str(i)].value) # ราคา
            text_out = "รหัสสินค้า " + text_id + " คือ " + product_name + " ราคา " + product_price + " บาท"
            break
        else:
            text_out = "ไม่พบรหัสสินค้านี้"
    wb.close()
    
    line_bot_api.reply_message(event.reply_token,
                             TextSendMessage(text=text_out))


def setup_handler():
    handler.add(MessageEvent, message=TextMessage)(handle_text_message)

if __name__ == "__main__":
    # โหลดตัวแปรจาก .env
    load_dotenv()
    channel_secret = os.getenv("CHANNEL_SECRET")
    channel_access_token = os.getenv("CHANNEL_ACCESS_TOKEN")
    
    init(channel_secret, channel_access_token)
    setup_handler()
          
    app.run()

